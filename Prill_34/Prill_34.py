# -*- coding: cp1251 -*-
from docx import Document
import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Side
from tkinter import*
from tkinter import filedialog
from tkinter import ttk
from tkinter.ttk import Combobox

god = ["2021-2022-2023","2022-2023-2024","2023-2024-2025","2024-2025-2026","2025-2026-2027","2026-2027-2028","2027-2028-2029","2028-2029-2030","2029-2030-2031","2030-2031-2032"]
godi = ['2022 год', '2023 год', '2024 год']
spis = []
t2spis = []
sohran = []
t2sohran = []
naz = []
sfera = dict()
subj = ['Выберите субъект']
with open("otdel.txt") as otd:
    for line in otd:
        sfera[line.strip().split()[0]] = line.strip().split()[1:]

def clicked():
    spis.append(filedialog.askopenfilename())
    naz = [n for i in spis for n in i.split('/')]
    lbl11 = Label(tab1, text = naz[-1], font = ('Arial Bold', 9))
    lbl11.place(relx = 0.02, rely = 0.27)

def soh():
    sohran.append(filedialog.askdirectory())
    lbl5 = Label(tab1, text = sohran[-1], font = ('Arial Bold', 9))
    lbl5.place(relx = 0.02, rely = 0.68)

def postv():
    combo['values'] = sorted(subj)

def t2clicked():
    t2spis.append(filedialog.askopenfilename())
    t2naz = [n for i in t2spis for n in i.split('/')]
    t2lbl6 = Label(tab2, text = t2naz[-1], font = ('Arial Bold', 9))
    t2lbl6.place(relx = 0.02, rely = 0.27)
    t2fil = op.reader.excel.load_workbook(filename = t2spis[-1])
    t2fil.active = 1
    i = 8
    while t2fil.active['A'+str(i)].value != None:
        subj.append(t2fil.active['A'+str(i)].value)
        i += 1

def t2soh():
    t2sohran.append(filedialog.askdirectory())
    t2lbl5 = Label(tab2, text = t2sohran[-1], font = ('Arial Bold', 9))
    t2lbl5.place(relx = 0.02, rely = 0.65)



        #Словарь таблиц и сфер из файла

def sform1():
    godi = [i + ' год' for i in combo1.get().split('-')]
    document = Document(spis[-1])
    documents = document.tables
    if txt.get() == 'все':
        naztabl = [int(tab) for sfer in sfera.keys() for tab in sfera[sfer]]
    else:
        naztabl = [int(i) for i in txt.get().split(', ')]
    wb = op.load_workbook('obsh.xlsx')
    
    for tablicc in naztabl:
        razdel = ''
        for i in sfera.keys():
            if str(tablicc) in sfera[i] and int(tablicc) <= len(documents)/2:
                razdel = i
                print(razdel)
                break
        if len(razdel) > 1:
            print('Таблица ' + str(tablicc))

            #шапка первой таблицы
            table = documents[(tablicc * 2) - 2]
            a = 1
            if tablicc == 1:
                a = 3
            naimen = table.cell(a,0).text
            print(naimen)
            stolbik = [column.cells[a + 2].text for column in table.columns if len(column.cells[a + 2].text) <= 9]

            # Список субъектов без шапки
            table = documents[(tablicc * 2) - 1]
            itog = dict()
            for row in table.rows:
                itog[row.cells[0].text] = [cell.text.replace(' ', '').replace(',', '.') for cell in row.cells if cell.text != row.cells[0].text]

                    # заполнение шаблона таблицы данными

            worksheet = wb[razdel]
            if razdel == 'Здравоохранение':
                worksheet['A1'] = r'Распределение межбюджетных трансфертов между субъектами Российской Федерации в сфере здравоохранения в ' + godi[0][:4] + '-' + godi[2][:4] + ' годах'
            if razdel == 'Культура':
                worksheet['A1'] = r'Распределение межбюджетных трансфертов между субъектами Российской Федерации в сфере культуры и туризма в ' + godi[0][:4] + '-' + godi[2][:4] + ' годах'
            if razdel == 'Спорт':
                worksheet['A1'] = r'Распределение межбюджетных трансфертов между субъектами Российской Федерации в сфере физической культуры и спорта в ' + godi[0][:4] + '-' + godi[2][:4] + ' годах'
            if razdel == 'Образование':
                worksheet['A1'] = r'Распределение межбюджетных трансфертов между субъектами Российской Федерации в сфере образования и молодежной политики в ' + godi[0][:4] + '-' + godi[2][:4] + ' годах'
            index = 2
            while True:
                if str(worksheet.cell(row = 3, column = index).value) != 'None':
                    index += 3 
                if str(worksheet.cell(row = 3, column = index).value) == 'None':
                    worksheet[get_column_letter(index) + '3'] = 'Таблица ' + str(tablicc)
                    worksheet[get_column_letter(index) + '4'] = naimen
                    worksheet[get_column_letter(index) + '7'] = godi[0]
                    worksheet[get_column_letter(index + 1) + '7'] = godi[1]
                    worksheet[get_column_letter(index + 2) + '7'] = godi[2]
                    peremen = 0
                    peremen2 = 0
                    for g in godi:
                        if g in stolbik:
                            for i in range(9, 104):
                                if worksheet['A' + str(i)].value in itog.keys():
                                    if itog[worksheet['A' + str(i)].value][peremen2] != '':
                                        worksheet[get_column_letter(index + peremen) + str(i)] = float(itog[worksheet['A' + str(i)].value][peremen2])
                            peremen2 += 1
                        peremen += 1

                            # генерируем суммы
                    for n in range(3):
                        worksheet[get_column_letter(index + n) + '8'] = '=SUM(' + get_column_letter(index + n) + '9:' + get_column_letter(index + n) + '26)'
                        worksheet[get_column_letter(index + n) + '27'] = '=SUM(' + get_column_letter(index + n) + '28:' + get_column_letter(index + n) + '38)'
                        worksheet[get_column_letter(index + n) + '39'] = '=SUM(' + get_column_letter(index + n) + '40:' + get_column_letter(index + n) + '47)'
                        worksheet[get_column_letter(index + n) + '48'] = '=SUM(' + get_column_letter(index + n) + '49:' + get_column_letter(index + n) + '55)'
                        worksheet[get_column_letter(index + n) + '56'] = '=SUM(' + get_column_letter(index + n) + '57:' + get_column_letter(index + n) + '70)'
                        worksheet[get_column_letter(index + n) + '71'] = '=SUM(' + get_column_letter(index + n) + '72:' + get_column_letter(index + n) + '77)'
                        worksheet[get_column_letter(index + n) + '78'] = '=SUM(' + get_column_letter(index + n) + '79:' + get_column_letter(index + n) + '88)'
                        worksheet[get_column_letter(index + n) + '89'] = '=SUM(' + get_column_letter(index + n) + '90:' + get_column_letter(index + n) + '102)'
                    print('сформирован')
                    break
    wb.save(sohran[-1] + '/Свод по субъектам ' + godi[0][:4] + '-' + godi[2][:4] + '.xlsx')
    print('Сохранено')
    window.destroy()

            # Генерация 2 таблицы
def sform2():
    obl = combo.get()
    t2fil = op.reader.excel.load_workbook(filename = t2spis[-1])
    lis = t2fil.sheetnames
    d = []

    for i in lis:
        t2ws = t2fil[i]
        m = 8
        while t2ws['A' + str(m)].value != obl:
            m += 1
        index = 2
        while t2ws.cell(row = 3, column = index).value != None:
            if t2ws.cell(row = m, column = index).value != None or t2ws.cell(row = m, column = index + 1).value != None or t2ws.cell(row = m, column = index + 2).value != None:
                d.append([t2ws.cell(row = 4, column = index).value, t2ws.cell(row = m, column = index).value, t2ws.cell(row = m, column = index + 1).value, t2ws.cell(row = m, column = index + 2).value, str(i)])
            index += 3
        if t2ws['B7'].value != None:
            godn = [t2ws['B7'].value, t2ws['C7'].value, t2ws['D7'].value]
    t2wb = op.load_workbook('subj.xlsx')
    t2wb2 = t2wb['Лист1']
    t2wb2['A1'] = 'Информация об объемах бюджетных трансфертов на ' + godn[0][0:4] + '-' + godn[2][0:4] + ' годы за счет средств федерального бюджета в отраслях социальной сферы'
    t2wb2['A3'] = obl
    t2wb2['C5'] = godn[0]
    t2wb2['D5'] = godn[1]
    t2wb2['E5'] = godn[2]
    k = 9
    kk = 9
    nom = 0
    nom2 = 0
    for i in d:
        if k == 9:
            nom += 1
            t2wb2['B'+str(kk)] = i[-1]
            t2wb2['B'+str(kk)].fill = PatternFill(fgColor="AFEEEE", fill_type = "solid")
            t2wb2['C'+str(kk)].fill = PatternFill(fgColor="AFEEEE", fill_type = "solid")
            t2wb2['D'+str(kk)].fill = PatternFill(fgColor="AFEEEE", fill_type = "solid")
            t2wb2['E'+str(kk)].fill = PatternFill(fgColor="AFEEEE", fill_type = "solid")
            #t2wb2.merge_cells('B' + str(kk) + ':E' + str(kk))
            t2wb2['B'+str(kk)].font = Font(size = 12, bold = True)
            t2wb2['A'+str(kk)] = nom
        k += 1
        if t2wb2['B'+str(kk)].value == i[-1]:
            nom2 += 1
            t2wb2['A'+str(k)] = str(nom) + '.' + str(nom2)
            t2wb2['B'+str(k)] = i[0]
            t2wb2['C'+str(k)] = i[1]
            t2wb2['D'+str(k)] = i[2]
            t2wb2['E'+str(k)] = i[3]
        else:
            kk = k
            nom += 1
            t2wb2['A'+str(kk)] = nom
            t2wb2['B'+str(kk)] = i[-1]
            t2wb2['B'+str(kk)].fill = PatternFill(fgColor="AFEEEE", fill_type = "solid")
            t2wb2['C'+str(kk)].fill = PatternFill(fgColor="AFEEEE", fill_type = "solid")
            t2wb2['D'+str(kk)].fill = PatternFill(fgColor="AFEEEE", fill_type = "solid")
            t2wb2['E'+str(kk)].fill = PatternFill(fgColor="AFEEEE", fill_type = "solid")
            #t2wb2.merge_cells('B' + str(kk) + ':E' + str(kk))
            t2wb2['B'+str(kk)].font = Font(size = 12, bold = True)
            k += 1
            nom2 = 1
            t2wb2['A'+str(k)] = str(nom) + '.' + str(nom2)
            t2wb2['B'+str(k)] = i[0]
            t2wb2['C'+str(k)] = i[1]
            t2wb2['D'+str(k)] = i[2]
            t2wb2['E'+str(k)] = i[3]
    t2wb2['C8'] = '=SUM('+ 'C9:C' + str(k) + ')'
    t2wb2['D8'] = '=SUM('+ 'D9:D' + str(k) + ')'
    t2wb2['E8'] = '=SUM('+ 'E9:E' + str(k) + ')'
    border = Border(left = Side(border_style = 'thin', color = '000000'), right = Side(border_style = 'thin', color = '000000'), top = Side(border_style = 'thin', color = '000000'), bottom = Side(border_style = 'thin', color = '000000'))
    rows = t2wb2['A9:E' + str(k)]
    for row in rows:
        for cell in row:
            cell.border = border
    if obl[-15:] == 'Санкт-Петербург':
        t2wb.save(t2sohran[-1] + '/Свод ' + obl[-15:] + '.xlsx')
    else:
        t2wb.save(t2sohran[-1] + '/Свод ' + obl + '.xlsx')
    print('Сохранено')
    window.destroy()









window = Tk()
window.title('Расоб 12')
window.geometry('480x300')
            #Делаем вкладки программы
tab_control = ttk.Notebook(window)
tab1 = ttk.Frame(tab_control)
tab2 = ttk.Frame(tab_control)
tab_control.add(tab1, text='Общая выборка')
tab_control.add(tab2, text='Выборка по округу')
            #Первая вкладка
lbl1 = Label(tab1, text = 'Выберите файл', font = ('Arial Bold', 14))
lbl1.place(relx = 0.02, rely = 0.05)
btn1 = Button(tab1, text = "Обзор", width =15, height = 1, bg = 'grey', fg = 'black', command=clicked)
btn1.place(relx = 0.02, rely = 0.15)

lbl2 = Label(tab1, text = 'Куда сохранить' +'\n' + ' итоговый файл?', font = ('Arial Bold', 14))
lbl2.place(relx = 0.01, rely = 0.4)
btn2 = Button(tab1, text = "Обзор", width =15, height = 1, bg = 'grey', fg = 'black', command=soh)
btn2.place(relx = 0.02, rely = 0.58)

btn6 = Button(tab1, text = "Сформировать файл", width =30, height = 2, bg = 'grey', fg = 'black', command=sform1)
btn6.place(relx = 0.534, rely = 0.85)

lbl3 = Label(tab1, text = 'Введите номера таблиц, \n или введите "все"', font = ('Arial Bold', 10))
lbl3.place(relx = 0.57, rely = 0.02)
txt = Entry(tab1, width = 38)
txt.place(relx = 0.5, rely = 0.15)

lbl4 = Label(tab1, text = 'Выберите период', font = ('Arial Bold', 10))
lbl4.place(relx = 0.57, rely = 0.5)
combo1 = Combobox(tab1, width=35) 
combo1['values'] = god
combo1.current(1)
combo1.place(relx = 0.5, rely = 0.6)


           #Вторая вкладка

t2lbl1 = Label(tab2, text = 'Выберите файл', font = ('Arial Bold', 14))
t2lbl1.place(relx = 0.02, rely = 0.05)
t2btn1 = Button(tab2, text = "Обзор", width =15, height = 1, bg = 'grey', fg = 'black', command=t2clicked)
t2btn1.place(relx = 0.02, rely = 0.15)

t2lbl2 = Label(tab2, text = 'Куда сохранить' +'\n' + ' итоговый файл?', font = ('Arial Bold', 14))
t2lbl2.place(relx = 0.01, rely = 0.35)
t2btn2 = Button(tab2, text = "Обзор", width =15, height = 1, bg = 'grey', fg = 'black', command=t2soh)
t2btn2.place(relx = 0.02, rely = 0.53)

t2btn3 = Button(tab2, text = "Сформировать файл", width =30, height = 2, bg = 'grey', fg = 'black', command=sform2)
t2btn3.place(relx = 0.534, rely = 0.85)

t2lbl4 = Label(tab2, text = 'Выберите округ', font = ('Arial Bold', 10))
t2lbl4.place(relx = 0.63, rely = 0.05)
combo = Combobox(tab2, postcommand=postv, width=35)
combo['values'] = subj
combo.current(0)
combo.place(relx = 0.5, rely = 0.15)


tab_control.pack(expand=1, fill='both')
window.mainloop()
